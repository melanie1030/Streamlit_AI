import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import silhouette_score
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
import seaborn as sns
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import statsmodels.api as sm
import xlsxwriter
from io import BytesIO
import openai
import json
import os

# 設置 OpenAI API key
openai.api_key = ""

# 設置頁面配置
st.set_page_config(
    page_title="商店銷售分析系統",
    page_icon="🏪",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 設置中文字體
plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei']
plt.rcParams['axes.unicode_minus'] = False

# 設置全局樣式
st.markdown("""
    <style>
    /* 主要內容區域背景 */
    .stApp {
        background: linear-gradient(-45deg, #E6F3FF, #E8FBFF, #FFF8E1);
        background-size: 400% 400%;
        animation: gradient 20s ease infinite;
        padding: 2rem;
    }
    
    @keyframes gradient {
        0% {
            background-position: 0% 50%;
        }
        50% {
            background-position: 100% 50%;
        }
        100% {
            background-position: 0% 50%;
        }
    }
    
    /* 側邊欄樣式 */
    section[data-testid="stSidebar"] {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.9), rgba(182, 251, 255, 0.2));
        border-right: 1px solid rgba(230, 230, 230, 0.5);
        box-shadow: 2px 0 15px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
    }

    section[data-testid="stSidebar"]:hover {
        box-shadow: 2px 0 20px rgba(135, 206, 235, 0.2);
    }

    /* 側邊欄內部元素 */
    section[data-testid="stSidebar"] .stMarkdown {
        background-color: transparent;
        transition: all 0.3s ease;
    }
    
    /* 標題樣式 */
    h1, h2, h3, h4, h5, h6 {
        color: #2E4053;
        font-weight: 700;
        padding: 0.5rem;
        margin-bottom: 1rem;
        background: linear-gradient(120deg, rgba(255, 255, 255, 0.9), rgba(182, 251, 255, 0.3));
        border-radius: 8px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.05);
        transition: all 0.3s ease;
    }

    h1:hover, h2:hover, h3:hover, h4:hover, h5:hover, h6:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 15px rgba(135, 206, 235, 0.3);
        background: linear-gradient(120deg, rgba(255, 255, 255, 0.95), rgba(173, 216, 230, 0.2));
    }

    /* 按鈕樣式 */
    .stButton > button {
        background: linear-gradient(120deg, #87CEEB, #ADD8E6);
        color: #2E4053;
        border: none;
        padding: 0.5rem 1rem;
        border-radius: 20px;
        font-weight: 500;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
        position: relative;
        overflow: hidden;
    }

    .stButton > button:hover {
        transform: translateY(-2px) scale(1.02);
        box-shadow: 0 4px 15px rgba(135, 206, 235, 0.4);
        background: linear-gradient(120deg, #ADD8E6, #87CEEB);
    }

    .stButton > button:active {
        transform: translateY(1px);
    }

    /* 添加彩虹邊框效果 */
    @keyframes borderGlow {
        0% {
            border-color: #87CEEB;
        }
        25% {
            border-color: #ADD8E6;
        }
        50% {
            border-color: #B6FBFF;
        }
        75% {
            border-color: #FFD700;
        }
        100% {
            border-color: #87CEEB;
        }
    }

    /* 其他樣式保持不變 */
    /* 數據框樣式 */
    .dataframe {
        background: rgba(255, 255, 255, 0.9);
        border: none !important;
        border-radius: 10px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        margin: 1rem 0;
        transition: all 0.3s ease;
    }

    .dataframe:hover {
        transform: translateY(-3px);
        box-shadow: 0 5px 15px rgba(255, 154, 139, 0.3);
    }

    .dataframe th {
        background-color: #faf9f6;
        color: #333333;
        font-weight: 600;
        padding: 12px !important;
        transition: all 0.3s ease;
    }

    .dataframe td {
        color: #333333;
        border: none !important;
        border-bottom: 1px solid rgba(182, 251, 255, 0.3) !important;
        padding: 12px !important;
        transition: all 0.2s ease;
    }

    .dataframe tr:hover {
        background-color: rgba(131, 164, 212, 0.1);
        transform: scale(1.01);
    }

    /* 圖表容器 */
    .stPlot {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.9), rgba(131, 164, 212, 0.1));
        border-radius: 10px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        padding: 1rem;
        margin: 1rem 0;
        transition: all 0.3s ease;
    }

    .stPlot:hover {
        transform: translateY(-3px) scale(1.01);
        box-shadow: 0 5px 15px rgba(255, 154, 139, 0.3);
    }

    /* 選擇器樣式 */
    .stSelectbox {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.9), rgba(255, 154, 139, 0.1));
        border-radius: 8px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
    }

    .stSelectbox:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 15px rgba(255, 106, 136, 0.2);
    }

    /* 指標卡片樣式 */
    div[data-testid="stMetricValue"] {
        background: linear-gradient(120deg, rgba(255, 255, 255, 0.9), rgba(182, 251, 255, 0.2));
        padding: 1rem;
        border-radius: 8px;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        margin: 0.5rem 0;
        transition: all 0.3s ease;
    }

    div[data-testid="stMetricValue"]:hover {
        transform: translateY(-2px) scale(1.02);
        box-shadow: 0 4px 15px rgba(131, 164, 212, 0.3);
    }

    /* Tab 樣式 */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.9), rgba(131, 164, 212, 0.1));
        border-radius: 10px;
        padding: 0.5rem;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
        transition: all 0.3s ease;
    }

    .stTabs [data-baseweb="tab-list"]:hover {
        box-shadow: 0 4px 15px rgba(255, 106, 136, 0.2);
    }

    .stTabs [data-baseweb="tab"] {
        background: linear-gradient(120deg, rgba(255, 255, 255, 0.9), rgba(182, 251, 255, 0.2));
        border-radius: 6px;
        padding: 0.5rem 1rem;
        color: #2E4053;
        transition: all 0.3s ease;
    }

    .stTabs [data-baseweb="tab"]:hover {
        transform: translateY(-1px);
        background: linear-gradient(120deg, rgba(255, 255, 255, 0.95), rgba(131, 164, 212, 0.2));
    }

    .stTabs [data-baseweb="tab-highlight"] {
        background: linear-gradient(120deg, #87CEEB, #ADD8E6);
        border-radius: 6px;
        transition: all 0.3s ease;
    }

    /* 添加全局動畫效果 */
    * {
        transition: background-color 0.3s ease;
    }

    /* 添加載入動畫 */
    @keyframes fadeIn {
        from {
            opacity: 0;
            transform: translateY(10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }

    .element-container {
        animation: fadeIn 0.5s ease-out;
    }
    </style>
""", unsafe_allow_html=True)

def create_figure_layout():
    return {
        'plot_bgcolor': 'white',
        'paper_bgcolor': 'white',
        'xaxis_title': "",
        'yaxis_title': "",
        'showlegend': False
    }

# 新增表格和文字樣式
st.markdown("""
    <style>
    /* 標題文字樣式 */
    h1, h2, h3, h4, h5, h6 {
        color: #333333;
        font-weight: 600;
    }

    /* 一般文字樣式 */
    p, span, div {
        color: #333333;
    }

    /* 表格樣式 */
    .dataframe {
        color: #333333;
        background-color: #ffffff;
    }

    .dataframe th {
        background-color: #faf9f6;
        color: #333333;
        font-weight: 600;
        padding: 8px;
    }

    .dataframe td {
        color: #333333;
        padding: 8px;
    }

    /* 表格hover效果 */
    .dataframe tr:hover {
        background-color: #faf9f6;
    }

    /* 數據標籤樣式 */
    .metric-label {
        color: #333333;
        font-weight: 500;
    }

    /* 圖表標題和軸標籤 */
    .plot-container text {
        color: #333333 !important;
        fill: #333333 !important;
    }

    /* Streamlit特定元素樣式 */
    .stMarkdown, .stText {
        color: #333333;
    }

    .st-bb {
        color: #333333;
    }

    .st-bw {
        color: #333333;
    }

    /* 連結顏色 */
    a {
        color: #666666;
    }

    a:hover {
        color: #333333;
    }

    /* 強調文字 */
    .highlight {
        color: #666666;
        font-weight: 600;
    }
    </style>
""", unsafe_allow_html=True)

def generate_profit_loss_statement(df):
    """生成損益表"""
    st.header("📊 損益分析")
    
    # 計算關鍵指標
    total_sales = df['Item_MRP'].sum()  # 使用商品價格總和作為銷售額
    avg_price = df['Item_MRP'].mean()
    total_items = len(df)
    
    # 創建兩列佈局
    col1, col2 = st.columns(2)
    
    with col1:
        st.metric(
            label="總銷售額",
            value=f"¥{total_sales:,.2f}",
            delta=None
        )
        
        st.metric(
            label="平均單價",
            value=f"¥{avg_price:,.2f}",
            delta=None
        )
    
    with col2:
        st.metric(
            label="商品數量",
            value=f"{total_items:,}",
            delta=None
        )
        
        if 'Item_Weight' in df.columns:
            total_weight = df['Item_Weight'].sum()
            st.metric(
                label="總重量",
                value=f"{total_weight:,.2f}kg",
                delta=None
            )
    
    # 按商品類型分析
    st.subheader("商品類型分析")
    if 'Item_Type' in df.columns:
        type_analysis = df.groupby('Item_Type').agg({
            'Item_MRP': ['sum', 'mean', 'count']
        }).round(2)
        
        type_analysis.columns = ['總銷售額', '平均價格', '商品數量']
        
        # 排序並顯示結果
        type_analysis = type_analysis.sort_values('總銷售額', ascending=False)
        
        # 使用plotly繪製條形圖
        fig = px.bar(
            type_analysis,
            y=type_analysis.index,
            x='總銷售額',
            title='各類型商品銷售額',
            labels={'總銷售額': '銷售額 (¥)', 'Item_Type': '商品類型'}
        )
        st.plotly_chart(fig)
        
        # 顯示詳細數據
        st.write("商品類型詳細數據：")
        st.dataframe(type_analysis.style.format({
            '總銷售額': '¥{:,.2f}',
            '平均價格': '¥{:,.2f}',
            '商品數量': '{:,.0f}'
        }))
    
    # 按商店分析
    st.subheader("商店分析")
    if 'Outlet_Type' in df.columns:
        store_analysis = df.groupby('Outlet_Type').agg({
            'Item_MRP': ['sum', 'mean', 'count']
        }).round(2)
        
        store_analysis.columns = ['總銷售額', '平均價格', '商品數量']
        
        # 排序並顯示結果
        store_analysis = store_analysis.sort_values('總銷售額', ascending=False)
        
        # 使用plotly繪製圓餅圖
        fig = px.pie(
            values=store_analysis['總銷售額'],
            names=store_analysis.index,
            title='各類型商店銷售額佔比'
        )
        st.plotly_chart(fig)
        
        # 顯示詳細數據
        st.write("商店類型詳細數據：")
        st.dataframe(store_analysis.style.format({
            '總銷售額': '¥{:,.2f}',
            '平均價格': '¥{:,.2f}',
            '商品數量': '{:,.0f}'
        }))
    
    # 時間趨勢分析
    st.subheader("時間趨勢分析")
    if 'Outlet_Establishment_Year' in df.columns:
        time_analysis = df.groupby('Outlet_Establishment_Year').agg({
            'Item_MRP': ['sum', 'mean', 'count']
        }).round(2)
        
        time_analysis.columns = ['總銷售額', '平均價格', '商品數量']
        
        # 使用plotly繪製折線圖
        fig = go.Figure()
        
        fig.add_trace(go.Scatter(
            x=time_analysis.index,
            y=time_analysis['總銷售額'],
            mode='lines+markers',
            name='總銷售額'
        ))
        
        fig.update_layout(
            title='銷售額時間趨勢',
            xaxis_title='年份',
            yaxis_title='銷售額 (¥)'
        )
        
        st.plotly_chart(fig)
        
        # 顯示詳細數據
        st.write("年度詳細數據：")
        st.dataframe(time_analysis.style.format({
            '總銷售額': '¥{:,.2f}',
            '平均價格': '¥{:,.2f}',
            '商品數量': '{:,.0f}'
        }))

def generate_customer_revenue_report(df):
    """生成客戶收入報表"""
    st.header("👥 客戶收入報表")
    
    # 按商店分析客戶收入
    customer_revenue = df.groupby('Outlet_Identifier').agg({
        'Item_MRP': ['sum', 'mean', 'count'],
        'Item_Weight': 'sum'
    }).round(2)
    
    customer_revenue.columns = ['總收入', '平均消費', '交易次數', '總重量']
    
    # 計算客戶貢獻度
    total_revenue = customer_revenue['總收入'].sum()
    customer_revenue['收入佔比'] = (customer_revenue['總收入'] / total_revenue * 100).round(2)
    
    # 排序並標記客戶等級
    customer_revenue = customer_revenue.sort_values('總收入', ascending=False)
    customer_revenue['客戶等級'] = pd.qcut(customer_revenue['總收入'], 
                                     q=3, 
                                     labels=['C級', 'B級', 'A級'])
    
    # 顯示客戶收入報表
    st.dataframe(customer_revenue.style.format({
        '總收入': '¥{:,.2f}',
        '平均消費': '¥{:,.2f}',
        '交易次數': '{:,.0f}',
        '總重量': '{:,.2f}kg',
        '收入佔比': '{:.1f}%'
    }))
    
    # 視覺化客戶分布
    fig = plt.figure(figsize=(10, 6))
    plt.pie(customer_revenue['總收入'], 
           labels=customer_revenue.index,
           autopct='%1.1f%%')
    plt.title('客戶收入分布')
    st.pyplot(fig)

def generate_balance_sheet(df):
    """生成資產負債表"""
    st.header("💰 資產負債表")
    
    # 計算資產項目
    total_inventory = (df['Item_MRP'] * df['Item_Weight']).sum()  # 假設庫存
    cash = df['Item_MRP'].sum() * 0.3  # 假設現金為總收入的30%
    accounts_receivable = df['Item_MRP'].sum() * 0.1  # 假設應收帳款為總收入的10%
    fixed_assets = df['Item_MRP'].sum() * 0.5  # 假設固定資產為總收入的50%
    
    total_assets = total_inventory + cash + accounts_receivable + fixed_assets
    
    # 計算負債項目
    accounts_payable = total_inventory * 0.4  # 假設應付帳款為庫存的40%
    short_term_debt = total_assets * 0.2  # 假設短期借款為總資產的20%
    long_term_debt = total_assets * 0.3  # 假設長期借款為總資產的30%
    
    total_liabilities = accounts_payable + short_term_debt + long_term_debt
    
    # 計算權益
    equity = total_assets - total_liabilities
    
    # 創建資產負債表
    bs_data = {
        '項目': ['資產', '  現金', '  應收帳款', '  存貨', '  固定資產', '資產總計',
                '負債', '  應付帳款', '  短期借款', '  長期借款', '負債總計',
                '權益', '負債及權益總計'],
        '金額': [None, cash, accounts_receivable, total_inventory, fixed_assets, total_assets,
                None, accounts_payable, short_term_debt, long_term_debt, total_liabilities,
                equity, total_liabilities + equity]
    }
    bs_df = pd.DataFrame(bs_data)
    
    # 顯示資產負債表
    st.dataframe(bs_df.style.format({
        '金額': '¥{:,.2f}'
    }))
    
    # 計算關鍵財務比率
    current_ratio = (cash + accounts_receivable + total_inventory) / (accounts_payable + short_term_debt)
    debt_ratio = total_liabilities / total_assets
    equity_ratio = equity / total_assets
    
    # 顯示關鍵比率
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("流動比率", f"{current_ratio:.2f}")
    with col2:
        st.metric("負債比率", f"{debt_ratio:.2%}")
    with col3:
        st.metric("權益比率", f"{equity_ratio:.2%}")

def generate_financial_ratios(df):
    """財務比率分析"""
    st.header("📈 財務比率分析")
    
    # 計算收入相關指標
    total_sales = df['Item_MRP'].sum()
    total_assets = total_sales * 2  # 假設總資產為銷售額的2倍
    total_equity = total_assets * 0.4  # 假設權益為總資產的40%
    net_income = total_sales * 0.1  # 假設淨利為銷售額的10%
    
    # 計算各項財務比率
    ratios = {
        '獲利能力比率': {
            '總資產報酬率 (ROA)': (net_income / total_assets) * 100,
            '股東權益報酬率 (ROE)': (net_income / total_equity) * 100,
            '淨利率': (net_income / total_sales) * 100
        },
        '營運效率比率': {
            '總資產週轉率': total_sales / total_assets,
            '存貨週轉率': total_sales / (total_sales * 0.2),  # 假設平均存貨為銷售額的20%
            '應收帳款週轉率': total_sales / (total_sales * 0.1)  # 假設平均應收帳款為銷售額的10%
        },
        '財務結構比率': {
            '負債比率': ((total_assets - total_equity) / total_assets) * 100,
            '權益比率': (total_equity / total_assets) * 100,
            '負債對權益比率': ((total_assets - total_equity) / total_equity) * 100
        }
    }
    
    # 顯示各類比率
    for category, category_ratios in ratios.items():
        st.subheader(category)
        ratio_df = pd.DataFrame({
            '比率': list(category_ratios.keys()),
            '數值': list(category_ratios.values())
        })
        st.dataframe(ratio_df.style.format({
            '數值': '{:.2f}'
        }))
        
        # 視覺化
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=list(category_ratios.keys()),
            y=list(category_ratios.values())
        ))
        fig.update_layout(title=category)
        st.plotly_chart(fig)

def generate_operational_metrics(df):
    """營運指標分析"""
    st.header("📊 營運指標分析")
    
    # 1. 銷售效率指標
    sales_metrics = {
        '平均單筆銷售額': df['Item_MRP'].mean(),
        '每平方米銷售額': df['Item_MRP'].sum() / len(df['Outlet_Identifier'].unique()),
        '每件商品平均利潤': (df['Item_MRP'] - df['Item_MRP'] * 0.7).mean()
    }
    
    # 2. 商品效率指標
    product_metrics = df.groupby('Item_Type').agg({
        'Item_MRP': ['mean', 'sum', 'count']
    }).round(2)
    product_metrics.columns = ['平均銷售額', '總銷售額', '銷售數量']
    
    # 3. 商店效率指標
    store_metrics = df.groupby('Outlet_Type').agg({
        'Item_MRP': ['mean', 'sum', 'count']
    }).round(2)
    store_metrics.columns = ['平均銷售額', '總銷售額', '銷售數量']
    
    # 顯示銷售效率指標
    st.subheader("銷售效率指標")
    sales_df = pd.DataFrame({
        '指標': list(sales_metrics.keys()),
        '數值': list(sales_metrics.values())
    })
    st.dataframe(sales_df.style.format({
        '數值': '¥{:,.2f}'
    }))
    
    # 顯示商品效率指標
    st.subheader("商品效率指標")
    st.dataframe(product_metrics.style.format({
        '平均銷售額': '¥{:,.2f}',
        '總銷售額': '¥{:,.2f}',
        '銷售數量': '{:,.0f}'
    }))
    
    # 顯示商店效率指標
    st.subheader("商店效率指標")
    st.dataframe(store_metrics.style.format({
        '平均銷售額': '¥{:,.2f}',
        '總銷售額': '¥{:,.2f}',
        '銷售數量': '{:,.0f}'
    }))
    
    # 視覺化商品和商店效率
    col1, col2 = st.columns(2)
    
    with col1:
        fig1 = plt.figure(figsize=(10, 6))
        product_metrics['總銷售額'].plot(kind='bar')
        plt.title('各類商品總銷售額')
        plt.xticks(rotation=45)
        st.pyplot(fig1)
    
    with col2:
        fig2 = plt.figure(figsize=(10, 6))
        store_metrics['總銷售額'].plot(kind='bar')
        plt.title('各類商店總銷售額')
        plt.xticks(rotation=45)
        st.pyplot(fig2)

def load_and_process_data(file):
    """載入並處理數據"""
    try:
        # 讀取上傳的文件
        df = pd.read_csv(file)
        
        # 檢查必要的列是否存在
        expected_columns = [
            'Item_Identifier', 'Item_Weight', 'Item_Fat_Content', 
            'Item_Visibility', 'Item_Type', 'Item_MRP', 
            'Outlet_Identifier', 'Outlet_Establishment_Year',
            'Outlet_Size', 'Outlet_Location_Type', 'Outlet_Type'
        ]
        
        missing_columns = [col for col in expected_columns if col not in df.columns]
        
        if missing_columns:
            st.error(f"數據缺少必要的列: {', '.join(missing_columns)}")
            st.info("""
            請確保您的CSV文件包含所有必要列。
            目前缺少的列已在上方列出。
            """)
            return None
        
        # 基礎數據清理
        # 1. 處理缺失值
        df['Item_Weight'].fillna(df['Item_Weight'].mean(), inplace=True)
        df['Outlet_Size'].fillna(df['Outlet_Size'].mode()[0], inplace=True)
        
        # 2. 標準化分類變量
        df['Item_Fat_Content'] = df['Item_Fat_Content'].str.lower()
        df['Item_Fat_Content'] = df['Item_Fat_Content'].replace({
            'reg': 'regular',
            'low fat': 'low_fat',
            'lowfat': 'low_fat',
            'lf': 'low_fat',
            'regular': 'regular'
        })
        
        # 3. 計算基本的統計特徵
        df['Store_Age'] = 2025 - df['Outlet_Establishment_Year']  # 計算商店年齡
        
        # 顯示數據基本信息
        st.success(f"""
        數據加載成功！
        - 總記錄數：{len(df):,} 筆
        - 商品種類數：{df['Item_Type'].nunique()} 種
        - 商店數量：{df['Outlet_Identifier'].nunique()} 家
        """)
        
        return df
        
    except pd.errors.EmptyDataError:
        st.error("上傳的文件是空的，請檢查文件內容。")
        return None
    except Exception as e:
        st.error(f"數據處理時發生錯誤：{str(e)}")
        st.info("請確保您上傳的是正確格式的CSV文件。")
        return None

def show_header_metrics(df):
    """顯示頂部關鍵指標"""
    # 計算關鍵指標
    total_items = len(df['Item_Identifier'].unique())
    total_stores = len(df['Outlet_Identifier'].unique())
    avg_item_mrp = df['Item_MRP'].mean()
    avg_visibility = df['Item_Visibility'].mean() * 100  # 轉換為百分比
    
    # 創建四個列來顯示指標
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            label="商品種類",
            value=f"{total_items:,}",
            help="獨特商品的總數量"
        )
    
    with col2:
        st.metric(
            label="商店數量", 
            value=f"{total_stores:,}",
            help="商店的總數量"
        )
    
    with col3:
        st.metric(
            label="平均商品價格",
            value=f"¥{avg_item_mrp:.2f}",
            help="商品的平均標價"
        )
    
    with col4:
        st.metric(
            label="平均商品能見度",
            value=f"{avg_visibility:.2f}%",
            help="商品的平均展示佔比"
        )

def analyze_products(df):
    """商品分析"""
    st.header("📦 商品分析")
    
    # 創建兩列佈局
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("商品類型分布")
        # 計算每種商品類型的統計數據
        type_stats = df.groupby('Item_Type').agg({
            'Item_Identifier': 'count',
            'Item_MRP': ['mean', 'min', 'max'],
            'Item_Weight': 'mean',
            'Item_Visibility': 'mean'
        }).round(2)
        
        # 重命名列
        type_stats.columns = [
            '商品數量',
            '平均價格',
            '最低價格',
            '最高價格',
            '平均重量',
            '平均能見度'
        ]
        
        # 顯示統計數據
        st.dataframe(type_stats.style.format({
            '平均價格': '¥{:.2f}',
            '最低價格': '¥{:.2f}',
            '最高價格': '¥{:.2f}',
            '平均重量': '{:.2f}kg',
            '平均能見度': '{:.2%}'
        }))
    
    with col2:
        st.subheader("商品價格分布")
        fig = px.box(df, 
                    x='Item_Type', 
                    y='Item_MRP',
                    title='各類型商品價格分布')
        fig.update_layout(
            xaxis_title="商品類型",
            yaxis_title="價格 (¥)",
            showlegend=False
        )
        st.plotly_chart(fig)
    
    # 商品脂肪含量分析
    st.subheader("商品脂肪含量分析")
    col3, col4 = st.columns(2)
    
    with col3:
        fat_content_dist = df['Item_Fat_Content'].value_counts()
        fig = px.pie(values=fat_content_dist.values,
                    names=fat_content_dist.index,
                    title='商品脂肪含量分布')
        st.plotly_chart(fig)
    
    with col4:
        avg_price_by_fat = df.groupby('Item_Fat_Content')['Item_MRP'].mean().round(2)
        fig = px.bar(x=avg_price_by_fat.index,
                    y=avg_price_by_fat.values,
                    title='不同脂肪含量商品的平均價格')
        fig.update_layout(
            xaxis_title="脂肪含量",
            yaxis_title="平均價格 (¥)"
        )
        st.plotly_chart(fig)
    
    # 商品能見度分析
    st.subheader("商品能見度分析")
    col5, col6 = st.columns(2)
    
    with col5:
        fig = px.histogram(df,
                          x='Item_Visibility',
                          title='商品能見度分布',
                          nbins=50)
        fig.update_layout(
            xaxis_title="能見度",
            yaxis_title="商品數量"
        )
        st.plotly_chart(fig)
    
    with col6:
        avg_visibility_by_type = df.groupby('Item_Type')['Item_Visibility'].mean().sort_values(ascending=False)
        fig = px.bar(x=avg_visibility_by_type.index,
                    y=avg_visibility_by_type.values,
                    title='各類型商品的平均能見度')
        fig.update_layout(
            xaxis_title="商品類型",
            yaxis_title="平均能見度",
            xaxis={'tickangle': 45}
        )
        st.plotly_chart(fig)

def analyze_stores(df):
    """商店分析"""
    st.header("🏪 商店分析")
    
    # 創建兩列佈局
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("商店類型分布")
        # 計算每種商店類型的統計數據
        store_stats = df.groupby('Outlet_Type').agg({
            'Item_Identifier': 'count',
            'Item_MRP': ['mean', 'min', 'max'],
            'Outlet_Identifier': 'nunique'
        }).round(2)
        
        # 重命名列
        store_stats.columns = [
            '商品數量',
            '平均商品價格',
            '最低商品價格',
            '最高商品價格',
            '商店數量'
        ]
        
        # 顯示統計數據
        st.dataframe(store_stats.style.format({
            '平均商品價格': '¥{:.2f}',
            '最低商品價格': '¥{:.2f}',
            '最高商品價格': '¥{:.2f}',
            '商品數量': '{:,.0f}',
            '商店數量': '{:,.0f}'
        }))
    
    with col2:
        st.subheader("商店規模分布")
        size_dist = df.groupby(['Outlet_Type', 'Outlet_Size']).size().unstack(fill_value=0)
        fig = px.bar(size_dist, 
                    title='不同類型商店的規模分布',
                    barmode='stack')
        fig.update_layout(
            xaxis_title="商店類型",
            yaxis_title="商店數量",
            showlegend=True,
            legend_title="商店規模"
        )
        st.plotly_chart(fig)
    
    # 商店位置分析
    st.subheader("商店位置分析")
    col3, col4 = st.columns(2)
    
    with col3:
        location_dist = df.groupby('Outlet_Location_Type')['Outlet_Identifier'].nunique()
        fig = px.pie(values=location_dist.values,
                    names=location_dist.index,
                    title='商店位置分布')
        st.plotly_chart(fig)
    
    with col4:
        avg_price_by_location = df.groupby('Outlet_Location_Type')['Item_MRP'].mean().round(2)
        fig = px.bar(x=avg_price_by_location.index,
                    y=avg_price_by_location.values,
                    title='不同位置商店的平均商品價格')
        fig.update_layout(
            xaxis_title="商店位置",
            yaxis_title="平均商品價格 (¥)"
        )
        st.plotly_chart(fig)
    
    # 商店年齡分析
    st.subheader("商店年齡分析")
    col5, col6 = st.columns(2)
    
    with col5:
        df['Store_Age'] = 2025 - df['Outlet_Establishment_Year']
        fig = px.histogram(df,
                          x='Store_Age',
                          title='商店年齡分布',
                          nbins=20)
        fig.update_layout(
            xaxis_title="商店年齡（年）",
            yaxis_title="商店數量"
        )
        st.plotly_chart(fig)
    
    with col6:
        avg_price_by_age = df.groupby('Store_Age')['Item_MRP'].mean().round(2)
        fig = px.line(x=avg_price_by_age.index,
                     y=avg_price_by_age.values,
                     title='商店年齡與平均商品價格的關係')
        fig.update_layout(
            xaxis_title="商店年齡（年）",
            yaxis_title="平均商品價格 (¥)"
        )
        st.plotly_chart(fig)

def analyze_trends(df):
    """趨勢分析"""
    st.header("📈 趨勢分析")
    
    # 創建兩列佈局
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("商店發展趨勢")
        # 按年份統計商店數量
        yearly_stores = df.groupby('Outlet_Establishment_Year')['Outlet_Identifier'].nunique()
        fig = px.line(x=yearly_stores.index,
                     y=yearly_stores.values,
                     title='商店數量發展趨勢')
        fig.update_layout(
            xaxis_title="成立年份",
            yaxis_title="商店數量"
        )
        st.plotly_chart(fig)
    
    with col2:
        st.subheader("價格趨勢")
        # 按年份分析平均商品價格
        yearly_prices = df.groupby('Outlet_Establishment_Year')['Item_MRP'].mean().round(2)
        fig = px.line(x=yearly_prices.index,
                     y=yearly_prices.values,
                     title='各年份商店的平均商品價格')
        fig.update_layout(
            xaxis_title="成立年份",
            yaxis_title="平均商品價格 (¥)"
        )
        st.plotly_chart(fig)
    
    # 商品類型趨勢
    st.subheader("商品類型趨勢")
    col3, col4 = st.columns(2)
    
    with col3:
        # 分析不同年份的商品類型分布
        type_by_year = pd.crosstab(df['Outlet_Establishment_Year'], df['Item_Type'])
        fig = px.area(type_by_year, title='商品類型隨時間的變化')
        fig.update_layout(
            xaxis_title="成立年份",
            yaxis_title="商品數量",
            showlegend=True,
            legend_title="商品類型"
        )
        st.plotly_chart(fig)
    
    with col4:
        # 分析不同商品類型的平均價格變化
        type_price_trend = df.pivot_table(
            values='Item_MRP',
            index='Outlet_Establishment_Year',
            columns='Item_Type',
            aggfunc='mean'
        ).round(2)
        
        fig = px.line(type_price_trend, 
                     title='不同商品類型的價格趨勢')
        fig.update_layout(
            xaxis_title="成立年份",
            yaxis_title="平均價格 (¥)",
            showlegend=True,
            legend_title="商品類型"
        )
        st.plotly_chart(fig)
    
    # 商店規模趨勢
    st.subheader("商店規模趨勢")
    col5, col6 = st.columns(2)
    
    with col5:
        # 分析不同年份的商店規模分布
        size_by_year = pd.crosstab(df['Outlet_Establishment_Year'], df['Outlet_Size'])
        fig = px.bar(size_by_year, 
                    title='商店規模隨時間的變化',
                    barmode='stack')
        fig.update_layout(
            xaxis_title="成立年份",
            yaxis_title="商店數量",
            showlegend=True,
            legend_title="商店規模"
        )
        st.plotly_chart(fig)
    
    with col6:
        # 分析不同商店規模的平均商品價格
        size_price_trend = df.pivot_table(
            values='Item_MRP',
            index='Outlet_Establishment_Year',
            columns='Outlet_Size',
            aggfunc='mean'
        ).round(2)
        
        fig = px.line(size_price_trend, 
                     title='不同規模商店的價格趨勢')
        fig.update_layout(
            xaxis_title="成立年份",
            yaxis_title="平均價格 (¥)",
            showlegend=True,
            legend_title="商店規模"
        )
        st.plotly_chart(fig)

def perform_advanced_analysis(df):
    """進階分析"""
    st.header("🔍 進階分析")
    
    # 創建分析選項
    analysis_type = st.selectbox(
        "選擇分析類型",
        ["客戶群集分析", "銷售預測模型", "關聯性分析"]
    )
    
    if analysis_type == "客戶群集分析":
        perform_customer_clustering(df)
    elif analysis_type == "銷售預測模型":
        perform_sales_prediction(df)
    elif analysis_type == "關聯性分析":
        perform_correlation_analysis(df)

def perform_customer_clustering(df):
    """客戶群集分析"""
    st.subheader("👥 客戶群集分析")
    
    # 選擇用於聚類的特徵
    features = ['Item_MRP', 'Item_Weight']
    
    # 創建特徵矩陣
    X = df[features].copy()
    
    # 標準化特徵
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    
    # 使用固定的聚類數量，避免計算最佳聚類數
    n_clusters = 3  # 使用固定的聚類數量
    
    # 執行K-means聚類
    kmeans = KMeans(n_clusters=n_clusters, random_state=42)
    df['Cluster'] = kmeans.fit_predict(X_scaled)
    
    # 視覺化結果
    fig = px.scatter(
        df, 
        x='Item_MRP', 
        y='Item_Weight',
        color='Cluster',
        title='客戶聚類分析結果',
        labels={
            'Item_MRP': '商品價格',
            'Item_Weight': '商品重量',
            'Cluster': '客戶群組'
        }
    )
    st.plotly_chart(fig)
    
    # 分析每個群組的特徵
    cluster_stats = df.groupby('Cluster').agg({
        'Item_MRP': ['mean', 'count'],
        'Item_Weight': ['mean', 'sum']
    }).round(2)
    
    cluster_stats.columns = ['平均價格', '商品數量', '平均重量', '總重量']
    st.write("各客戶群組特徵：")
    st.dataframe(cluster_stats.style.format({
        '平均價格': '¥{:.2f}',
        '平均重量': '{:.2f}kg',
        '商品數量': '{:.0f}',
        '總重量': '{:.2f}kg'
    }))
    
    # 視覺化和描述每個群組
    for i in range(n_clusters):
        stats = cluster_stats.iloc[i]
        with st.container():
            st.write(f"""群組 {i}：
            - 商品數量：{stats['商品數量']:.0f}
            - 平均商品價格：¥{stats['平均價格']:.2f}
            - 平均商品重量：{stats['平均重量']:.2f}kg
            - 總商品重量：{stats['總重量']:.2f}kg
            """)
    
    # 生成群組建議
    st.subheader("群組行銷建議")
    
    recommendations = {
        '高價值群組': """
        - 提供高端商品推薦
        - 開發會員專屬服務
        - 提供個性化購物體驗
        """,
        '中價值群組': """
        - 提供性價比商品推薦
        - 開發促銷活動
        - 提供會員積分優惠
        """,
        '價格敏感群組': """
        - 提供特價商品推薦
        - 開發限時折扣活動
        - 提供優惠券
        """
    }
    
    for group, rec in recommendations.items():
        st.write(f"**{group}**")
        st.write(rec)

def perform_sales_prediction(df):
    """銷售預測模型"""
    st.subheader("🔮 銷售預測模型")
    
    # 準備特徵
    features = ['Item_MRP']
    if 'Item_Weight' in df.columns:
        features.append('Item_Weight')
    if 'Store_Age' in df.columns:
        features.append('Store_Age')
    
    # 準備數據
    X = df[features]
    y = df['Item_MRP']
    
    # 分割訓練集和測試集
    from sklearn.model_selection import train_test_split
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    
    # 訓練隨機森林模型
    model = RandomForestRegressor(n_estimators=100, random_state=42)
    model.fit(X_train, y_train)
    
    # 特徵重要性
    feature_importance = pd.DataFrame({
        '特徵': features,
        '重要性': model.feature_importances_
    }).sort_values('重要性', ascending=False)
    
    st.write("特徵重要性：")
    fig = px.bar(feature_importance,
                x='特徵',
                y='重要性',
                title='特徵重要性分析')
    st.plotly_chart(fig)
    
    # 預測結果
    y_pred = model.predict(X_test)
    from sklearn.metrics import r2_score, mean_squared_error
    r2 = r2_score(y_test, y_pred)
    rmse = np.sqrt(mean_squared_error(y_test, y_pred))
    
    col1, col2 = st.columns(2)
    with col1:
        st.metric("R² 分數", f"{r2:.3f}")
    with col2:
        st.metric("RMSE", f"{rmse:.2f}")
    
    # 預測vs實際值散點圖
    prediction_df = pd.DataFrame({
        '實際值': y_test,
        '預測值': y_pred
    })
    
    fig = px.scatter(prediction_df,
                    x='實際值',
                    y='預測值',
                    title='預測值 vs 實際值')
    fig.add_trace(go.Scatter(x=[y_test.min(), y_test.max()],
                            y=[y_test.min(), y_test.max()],
                            mode='lines',
                            name='完美預測線'))
    st.plotly_chart(fig)

def perform_correlation_analysis(df):
    """執行相關性分析"""
    st.subheader("🔄 相關性分析")
    
    # 選擇數值型特徵
    numeric_features = df.select_dtypes(include=['float64', 'int64']).columns
    
    # 使用SimpleImputer處理缺失值
    from sklearn.impute import SimpleImputer
    imputer = SimpleImputer(strategy='mean')
    X = df[numeric_features].copy()
    X_imputed = pd.DataFrame(imputer.fit_transform(X), columns=X.columns)
    
    # 計算相關係數矩陣
    corr_matrix = X_imputed.corr()
    
    # 繪製熱力圖
    fig = px.imshow(
        corr_matrix,
        labels=dict(color="相關係數"),
        x=corr_matrix.columns,
        y=corr_matrix.columns,
        color_continuous_scale="RdBu",
        title="特徵相關性熱力圖"
    )
    fig.update_layout(width=800, height=800)
    st.plotly_chart(fig)
    
    # 執行PCA分析
    st.subheader("主成分分析 (PCA)")
    
    # 標準化數據
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_imputed)
    
    # 執行PCA
    pca = PCA()
    X_pca = pca.fit_transform(X_scaled)
    
    # 計算解釋方差比
    explained_variance_ratio = pca.explained_variance_ratio_
    cumulative_variance_ratio = np.cumsum(explained_variance_ratio)
    
    # 繪製解釋方差比圖
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=[f'PC{i+1}' for i in range(len(explained_variance_ratio))],
        y=explained_variance_ratio,
        name='解釋方差比'
    ))
    fig.add_trace(go.Scatter(
        x=[f'PC{i+1}' for i in range(len(cumulative_variance_ratio))],
        y=cumulative_variance_ratio,
        name='累積解釋方差比',
        line=dict(color='red')
    ))
    fig.update_layout(title='主成分解釋方差比',
                     xaxis_title='主成分',
                     yaxis_title='解釋方差比',
                     barmode='group')
    st.plotly_chart(fig)
    
    # 顯示主成分載荷量
    loadings = pd.DataFrame(
        pca.components_.T,
        columns=[f'PC{i+1}' for i in range(len(pca.components_))],
        index=numeric_features
    )
    
    st.write("### 主成分載荷量")
    st.dataframe(loadings.style.format("{:.3f}"))
    
    # 分析和解釋主要相關性
    st.write("### 主要相關性分析")
    
    # 找出強相關的特徵對
    strong_correlations = []
    for i in range(len(numeric_features)):
        for j in range(i+1, len(numeric_features)):
            corr = corr_matrix.iloc[i,j]
            if abs(corr) > 0.5:  # 設定相關係數閾值
                strong_correlations.append({
                    'feature1': numeric_features[i],
                    'feature2': numeric_features[j],
                    'correlation': corr
                })
    
    # 按相關係數絕對值排序
    strong_correlations.sort(key=lambda x: abs(x['correlation']), reverse=True)
    
    # 顯示強相關特徵對
    if strong_correlations:
        for corr in strong_correlations:
            correlation_type = "正相關" if corr['correlation'] > 0 else "負相關"
            st.write(f"**{corr['feature1']} 和 {corr['feature2']}**")
            st.write(f"- 相關係數: {corr['correlation']:.3f} ({correlation_type})")
            
            # 繪製散點圖
            fig = px.scatter(df,
                            x=corr['feature1'],
                            y=corr['feature2'],
                            title=f"{corr['feature1']} vs {corr['feature2']}",
                            trendline="ols")
            st.plotly_chart(fig)
            
            # 生成業務建議
            if abs(corr['correlation']) > 0.7:
                st.write("💡 **強相關性建議：**")
                if corr['correlation'] > 0:
                    st.write(f"- 考慮將{corr['feature1']}和{corr['feature2']}作為組合指標")
                    st.write(f"- 可以通過提升{corr['feature1']}來帶動{corr['feature2']}的增長")
                else:
                    st.write(f"- 注意{corr['feature1']}和{corr['feature2']}之間的權衡關係")
                    st.write(f"- 需要在兩者之間找到最佳平衡點")
    else:
        st.write("未發現顯著的特徵相關性（相關係數絕對值 > 0.5）")

def perform_time_series_analysis(df):
    """時間序列分析"""
    st.header("📅 時間序列分析")
    
    # 檢查是否有日期列
    date_columns = df.select_dtypes(include=['datetime64']).columns
    if len(date_columns) == 0:
        st.warning("未找到日期列。請確保數據中包含日期信息。")
        return
    
    # 選擇日期列
    date_column = st.selectbox("選擇日期列", date_columns)
    
    # 設置日期索引
    df_ts = df.copy()
    df_ts.set_index(date_column, inplace=True)
    
    # 時間序列分解
    try:
        from statsmodels.tsa.seasonal import seasonal_decompose
        
        # 重採樣到日期級別並計算平均銷售額
        daily_sales = df_ts['Item_MRP'].resample('D').mean()
        
        # 填充缺失值
        daily_sales = daily_sales.fillna(daily_sales.mean())
        
        # 執行時間序列分解
        decomposition = seasonal_decompose(daily_sales, period=30)
        
        # 創建四個子圖
        fig = make_subplots(rows=4, cols=1,
                           subplot_titles=('原始數據', '趨勢', '季節性', '殘差'))
        
        # 添加原始數據
        fig.add_trace(
            go.Scatter(x=daily_sales.index, y=daily_sales.values,
                      name='原始數據'),
            row=1, col=1
        )
        
        # 添加趨勢
        fig.add_trace(
            go.Scatter(x=daily_sales.index, y=decomposition.trend,
                      name='趨勢'),
            row=2, col=1
        )
        
        # 添加季節性
        fig.add_trace(
            go.Scatter(x=daily_sales.index, y=decomposition.seasonal,
                      name='季節性'),
            row=3, col=1
        )
        
        # 添加殘差
        fig.add_trace(
            go.Scatter(x=daily_sales.index, y=decomposition.resid,
                      name='殘差'),
            row=4, col=1
        )
        
        # 更新布局
        fig.update_layout(height=800, title_text="時間序列分解")
        st.plotly_chart(fig)
        
        # 移動平均分析
        st.subheader("移動平均分析")
        
        # 選擇移動平均窗口大小
        window_size = st.slider("選擇移動平均窗口大小（天）", 
                              min_value=1, 
                              max_value=30, 
                              value=7)
        
        # 計算移動平均
        rolling_mean = daily_sales.rolling(window=window_size).mean()
        
        # 繪製移動平均圖
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=daily_sales.index, 
                                y=daily_sales.values,
                                name='原始數據'))
        fig.add_trace(go.Scatter(x=rolling_mean.index,
                                y=rolling_mean.values,
                                name=f'{window_size}天移動平均'))
        fig.update_layout(title=f'{window_size}天移動平均分析',
                         xaxis_title='日期',
                         yaxis_title='銷售額')
        st.plotly_chart(fig)
        
        # 自相關分析
        st.subheader("自相關分析")
        
        from statsmodels.tsa.stattools import acf, pacf
        
        # 計算自相關係數
        lag_acf = acf(daily_sales.dropna(), nlags=40)
        lag_pacf = pacf(daily_sales.dropna(), nlags=40)
        
        # 創建兩個子圖
        fig = make_subplots(rows=2, cols=1,
                           subplot_titles=('自相關函數 (ACF)',
                                         '偏自相關函數 (PACF)'))
        
        # 添加 ACF
        fig.add_trace(
            go.Scatter(x=list(range(len(lag_acf))),
                      y=lag_acf,
                      mode='lines+markers',
                      name='ACF'),
            row=1, col=1
        )
        
        # 添加 PACF
        fig.add_trace(
            go.Scatter(x=list(range(len(lag_pacf))),
                      y=lag_pacf,
                      mode='lines+markers',
                      name='PACF'),
            row=2, col=1
        )
        
        # 更新布局
        fig.update_layout(height=600,
                         showlegend=False,
                         title_text="自相關分析")
        st.plotly_chart(fig)
        
        # 銷售預測
        st.subheader("銷售預測")
        
        from statsmodels.tsa.arima.model import ARIMA
        
        # 訓練ARIMA模型
        model = ARIMA(daily_sales, order=(1,1,1))
        results = model.fit()
        
        # 預測未來30天
        forecast = results.forecast(steps=30)
        
        # 創建預測圖
        fig = go.Figure()
        
        # 添加歷史數據
        fig.add_trace(go.Scatter(x=daily_sales.index,
                                y=daily_sales.values,
                                name='歷史數據'))
        
        # 添加預測數據
        fig.add_trace(go.Scatter(x=pd.date_range(start=daily_sales.index[-1],
                                                periods=31)[1:],
                                y=forecast,
                                name='預測',
                                line=dict(dash='dash')))
        
        # 更新布局
        fig.update_layout(title='未來30天銷售預測',
                         xaxis_title='日期',
                         yaxis_title='銷售額')
        st.plotly_chart(fig)
        
        # 顯示預測結果表格
        forecast_df = pd.DataFrame({
            '日期': pd.date_range(start=daily_sales.index[-1],
                               periods=31)[1:],
            '預測銷售額': forecast.round(2)
        })
        
        st.write("未來30天預測結果：")
        st.dataframe(forecast_df.style.format({
            '預測銷售額': '¥{:,.2f}'
        }))
        
    except Exception as e:
        st.error(f"進行時間序列分析時發生錯誤: {str(e)}")

def perform_price_analysis(df):
    """價格分析"""
    st.header("💰 價格分析")
    
    if 'Item_MRP' not in df.columns:
        st.warning("未找到價格列（Item_MRP）。請確保數據中包含價格信息。")
        return
    
    # 基本價格統計
    st.subheader("基本價格統計")
    price_stats = df['Item_MRP'].describe()
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("平均價格", f"¥{price_stats['mean']:.2f}")
    with col2:
        st.metric("最高價格", f"¥{price_stats['max']:.2f}")
    with col3:
        st.metric("最低價格", f"¥{price_stats['min']:.2f}")
    with col4:
        st.metric("價格標準差", f"¥{price_stats['std']:.2f}")
    
    # 價格分布分析
    st.subheader("價格分布分析")
    
    # 創建價格區間
    price_bins = pd.qcut(df['Item_MRP'], q=10, labels=[f'第{i+1}分位' for i in range(10)])
    price_distribution = df.groupby(price_bins).agg({
        'Item_MRP': ['count', 'mean'],
        'Item_Weight': 'sum'
    })
    
    price_distribution.columns = ['商品數量', '平均價格', '總重量']
    
    # 顯示價格分布表格
    st.write("價格分布統計：")
    st.dataframe(price_distribution.style.format({
        '商品數量': '{:,d}',
        '平均價格': '¥{:,.2f}',
        '總重量': '{:,.2f}kg'
    }))
    
    # 價格分布直方圖
    fig = px.histogram(df,
                      x='Item_MRP',
                      nbins=30,
                      title='價格分布直方圖',
                      labels={'Item_MRP': '價格', 'count': '商品數量'})
    st.plotly_chart(fig)
    
    # 價格與銷售額的關係
    st.subheader("價格與銷售額關係分析")
    
    # 散點圖
    fig = px.scatter(df,
                    x='Item_MRP',
                    y='Item_Weight',
                    title='價格與銷售額關係',
                    labels={'Item_MRP': '價格',
                           'Item_Weight': '銷售額'})
    st.plotly_chart(fig)
    
    # 計算價格彈性
    if 'Item_Weight' in df.columns:
        # 計算平均價格和平均銷量
        avg_price = df['Item_MRP'].mean()
        avg_sales = df['Item_Weight'].mean()
        
        # 計算價格變化率和銷量變化率
        price_pct_change = (df['Item_MRP'] - avg_price) / avg_price
        sales_pct_change = (df['Item_Weight'] - avg_sales) / avg_sales
        
        # 計算價格彈性
        price_elasticity = sales_pct_change.mean() / price_pct_change.mean()
        
        st.metric("價格彈性", f"{abs(price_elasticity):.2f}",
                 help="價格彈性表示價格變動對銷售量的影響程度。\n"
                      "數值越大表示價格變動對銷售量的影響越大。")
    
    # 按商品類型的價格分析
    if 'Item_Type' in df.columns:
        st.subheader("商品類型價格分析")
        
        # 計算每種商品類型的價格統計
        type_price_stats = df.groupby('Item_Type').agg({
            'Item_MRP': ['mean', 'min', 'max', 'std'],
            'Item_Weight': 'sum'
        }).round(2)
        
        type_price_stats.columns = ['平均價格', '最低價格', '最高價格', '價格標準差', '總重量']
        
        # 排序並顯示
        type_price_stats = type_price_stats.sort_values('平均價格', ascending=False)
        
        st.write("商品類型價格統計：")
        st.dataframe(type_price_stats.style.format({
            '平均價格': '¥{:,.2f}',
            '最低價格': '¥{:,.2f}',
            '最高價格': '¥{:,.2f}',
            '價格標準差': '¥{:,.2f}',
            '總重量': '{:,.2f}kg'
        }))
        
        # 箱形圖
        fig = px.box(df,
                    x='Item_Type',
                    y='Item_MRP',
                    title='各類型商品價格分布',
                    labels={'Item_Type': '商品類型',
                           'Item_MRP': '價格'})
        fig.update_xaxes(tickangle=45)
        st.plotly_chart(fig)
    
    # 價格優化建議
    st.subheader("價格優化建議")
    
    # 分析高利潤商品
    if all(col in df.columns for col in ['Item_MRP', 'Item_Weight']):
        df['Profit_Margin'] = (df['Item_Weight'] - df['Item_MRP']) / df['Item_MRP']
        
        high_profit_items = df[df['Profit_Margin'] > df['Profit_Margin'].quantile(0.75)]
        
        st.write("高利潤商品分析：")
        high_profit_stats = high_profit_items.groupby('Item_Type').agg({
            'Item_MRP': 'mean',
            'Profit_Margin': 'mean',
            'Item_Identifier': 'count'
        }).round(3)
        
        high_profit_stats.columns = ['平均價格', '平均利潤率', '商品數量']
        st.dataframe(high_profit_stats.style.format({
            '平均價格': '¥{:,.2f}',
            '平均利潤率': '{:.1%}',
            '商品數量': '{:,d}'
        }))
        
        # 生成價格優化建議
        st.write("價格優化建議：")
        
        # 基於價格彈性的建議
        if abs(price_elasticity) > 1:
            st.info("📊 市場對價格較為敏感，建議：\n"
                   "1. 考慮實施差異化定價策略\n"
                   "2. 進行小幅度價格調整測試\n"
                   "3. 關注競爭對手的價格變動")
        else:
            st.info("📊 市場對價格相對不敏感，建議：\n"
                   "1. 可以考慮提高高利潤商品的價格\n"
                   "2. 重點關注產品質量和品牌建設\n"
                   "3. 開發高端市場細分")
        
        # 基於利潤率的建議
        high_profit_categories = high_profit_stats[high_profit_stats['平均利潤率'] > 0.2].index.tolist()
        if high_profit_categories:
            st.info(f"💡 以下類別具有較高利潤率，建議增加庫存和促銷力度：\n" + 
                   "\n".join([f"- {cat}" for cat in high_profit_categories]))

def perform_customer_segmentation(df):
    """客戶分群分析"""
    st.header("👥 客戶分群分析")
    
    # 檢查必要的列
    required_columns = ['Customer_ID', 'Item_Weight']
    if not all(col in df.columns for col in required_columns):
        st.warning("未找到客戶ID或銷售額列。請確保數據中包含客戶信息。")
        return
    
    # 計算客戶指標
    customer_metrics = df.groupby('Customer_ID').agg({
        'Item_Weight': ['sum', 'mean', 'count'],
        'Item_MRP': 'mean'
    })
    
    customer_metrics.columns = ['總消費額', '平均消費額', '購買次數', '平均商品價格']
    
    # 添加RFM指標
    if 'Transaction_Date' in df.columns:
        latest_date = df['Transaction_Date'].max()
        customer_metrics['最近購買間隔'] = (
            latest_date - df.groupby('Customer_ID')['Transaction_Date'].max()
        ).dt.days
    
    # 標準化數據
    scaler = StandardScaler()
    features_for_clustering = ['總消費額', '購買次數', '平均商品價格']
    if '最近購買間隔' in customer_metrics.columns:
        features_for_clustering.append('最近購買間隔')
    
    X = scaler.fit_transform(customer_metrics[features_for_clustering])
    
    # 使用K-means進行客戶分群
    n_clusters = st.slider("選擇客戶群數量", 2, 6, 3)
    kmeans = KMeans(n_clusters=n_clusters, random_state=42)
    customer_metrics['Customer_Segment'] = kmeans.fit_predict(X)
    
    # 分析各群特徵
    segment_analysis = customer_metrics.groupby('Customer_Segment').agg({
        '總消費額': 'mean',
        '平均消費額': 'mean',
        '購買次數': 'mean',
        '平均商品價格': 'mean',
        'Customer_ID': 'count'
    })
    
    segment_analysis.columns = ['平均總消費額', '平均單次消費額', '平均購買次數', 
                              '平均商品價格', '客戶數量']
    
    # 顯示分群結果
    st.subheader("客戶群特徵分析")
    st.dataframe(segment_analysis.style.format({
        '平均總消費額': '¥{:,.2f}',
        '平均單次消費額': '¥{:,.2f}',
        '平均購買次數': '{:,.1f}',
        '平均商品價格': '¥{:,.2f}',
        '客戶數量': '{:,d}'
    }))
    
    # 視覺化分群結果
    st.subheader("客戶群分布視覺化")
    
    # 使用PCA降維以便視覺化
    pca = PCA(n_components=2)
    X_pca = pca.fit_transform(X)
    
    # 創建視覺化數據框
    viz_df = pd.DataFrame(X_pca, columns=['PC1', 'PC2'])
    viz_df['Segment'] = customer_metrics['Customer_Segment']
    
    # 散點圖
    fig = px.scatter(viz_df,
                    x='PC1',
                    y='PC2',
                    color='Segment',
                    title='客戶群分布',
                    labels={'PC1': '主成分1', 'PC2': '主成分2'})
    st.plotly_chart(fig)
    
    # 客戶價值分析
    st.subheader("客戶價值分析")
    
    # 計算客戶終身價值 (CLV)
    customer_metrics['Customer_Lifetime_Value'] = (
        customer_metrics['總消費額'] * 
        (customer_metrics['購買次數'] > customer_metrics['購買次數'].mean()).astype(int) * 
        1.5  # 假設高頻客戶的未來價值更高
    )
    
    # 計算各群的平均CLV
    clv_by_segment = customer_metrics.groupby('Customer_Segment')[
        'Customer_Lifetime_Value'
    ].mean().sort_values(ascending=False)
    
    # 顯示CLV分析
    fig = px.bar(x=clv_by_segment.index,
                y=clv_by_segment.values,
                title='各客戶群平均終身價值',
                labels={'x': '客戶群', 'y': '平均終身價值'})
    st.plotly_chart(fig)
    
    # 客戶群特徵描述
    st.subheader("客戶群特徵描述")
    
    for segment in range(n_clusters):
        segment_stats = segment_analysis.loc[segment]
        
        # 判斷客戶群特徵
        if segment_stats['平均總消費額'] > segment_analysis['平均總消費額'].mean():
            value_level = "高價值"
        else:
            value_level = "一般價值"
        
        if segment_stats['平均購買次數'] > segment_analysis['平均購買次數'].mean():
            frequency_level = "高頻"
        else:
            frequency_level = "低頻"
        
        with st.container():
            # 生成描述
            st.write(f"""### 客戶群 {segment} ({value_level}, {frequency_level}客戶)
            - 客戶數量: {segment_stats['客戶數量']:,d} 人
            - 平均總消費: ¥{segment_stats['平均總消費額']:,.2f}
            - 平均購買次數: {segment_stats['平均購買次數']:.1f} 次
            - 平均單次消費: ¥{segment_stats['平均單次消費額']:,.2f}
            """)
            
            # 為每個群提供營銷建議
            if value_level == "高價值" and frequency_level == "高頻":
                recommendations = [
                    "提供VIP專屬服務和優惠",
                    "開發高端產品線",
                    "建立忠誠度計劃"
                ]
            elif value_level == "高價值" and frequency_level == "低頻":
                recommendations = [
                    "增加互動頻率",
                    "提供個性化服務",
                    "發送專屬優惠"
                ]
            elif value_level == "一般價值" and frequency_level == "高頻":
                recommendations = [
                    "提升單次消費額",
                    "推薦相關產品",
                    "提供產品組合優惠"
                ]
            else:
                recommendations = [
                    "提供入門級產品",
                    "發送促銷優惠",
                    "提高品牌認知"
                ]
            
            st.write("**營銷建議：**")
            for rec in recommendations:
                st.write(f"- {rec}")

def generate_excel_report(df, customer_metrics=None, segment_analysis=None):
    """生成專業的Excel分析報表"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image
    from io import BytesIO
    import matplotlib.pyplot as plt
    import seaborn as sns
    from datetime import datetime
    
    wb = Workbook()
    
    # 定義專業的顏色方案和樣式
    COLORS = {
        'header_bg': "1F4E78",  # 深藍色表頭
        'subheader_bg': "2F75B5",  # 中藍色副表頭
        'highlight': "BDD7EE",  # 淺藍色強調
        'total_row': "DDEBF7"  # 總計行背景
    }
    
    # 設置通用樣式
    header_fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type="solid")
    subheader_fill = PatternFill(start_color=COLORS['subheader_bg'], end_color=COLORS['subheader_bg'], fill_type="solid")
    highlight_fill = PatternFill(start_color=COLORS['highlight'], end_color=COLORS['highlight'], fill_type="solid")
    total_fill = PatternFill(start_color=COLORS['total_row'], end_color=COLORS['total_row'], fill_type="solid")
    
    header_font = Font(name='Arial', size=11, color="FFFFFF", bold=True)
    normal_font = Font(name='Arial', size=10)
    title_font = Font(name='Arial', size=14, bold=True)
    
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    # 1. 封面頁面
    ws_cover = wb.active
    ws_cover.title = "封面"
    
    # 設置封面
    ws_cover['A1'].font = Font(name='Arial', size=24, bold=True)
    ws_cover['A1'] = "銷售分析報告"
    ws_cover.merge_cells('A1:E1')
    ws_cover['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_cover['A3'].font = Font(name='Arial', size=12)
    ws_cover['A3'] = f"報告生成日期：{datetime.now().strftime('%Y年%m月%d日')}"
    
    ws_cover['A5'].font = Font(name='Arial', size=12)
    ws_cover['A5'] = "報告內容："
    ws_cover['A6'] = "1. 經營概況摘要"
    ws_cover['A7'] = "2. 銷售績效分析"
    ws_cover['A8'] = "3. 商品類別分析"
    if customer_metrics is not None:
        ws_cover['A9'] = "4. 客戶群分析"
    
    # 2. 經營概況摘要
    ws_summary = wb.create_sheet("經營概況摘要")
    
    # 設置標題
    ws_summary['A1'] = "經營概況摘要報告"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:E1')
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    
    # KPI指標
    headers = ["關鍵績效指標", "數值", "同比變化", "說明"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    # 計算KPI
    total_items = len(df['Item_Identifier'].unique())
    total_stores = len(df['Outlet_Identifier'].unique())
    avg_item_mrp = df['Item_MRP'].mean()
    avg_visibility = df['Item_Visibility'].mean() * 100  # 轉換為百分比
    
    kpis = [
        ["總商品數", total_items, "-", "所有商品的總數量"],
        ["總商店數", total_stores, "-", "所有商店的總數量"],
        ["平均商品價格", avg_item_mrp, "-", "商品的平均標價"],
        ["平均商品能見度", avg_visibility, "-", "商品的平均展示佔比"],
    ]
    
    for row, kpi in enumerate(kpis, 4):
        for col, value in enumerate(kpi, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
            if col == 2:  # 數值列
                cell.number_format = '$#,##0.00'
    
    # 3. 銷售績效分析
    ws_sales = wb.create_sheet("銷售績效分析")
    
    ws_sales['A1'] = "銷售績效分析報告"
    ws_sales['A1'].font = title_font
    ws_sales.merge_cells('A1:E1')
    ws_sales['A1'].alignment = Alignment(horizontal='center')
    
    # 門店銷售分析
    outlet_sales = df.groupby('Outlet_Identifier').agg({
        'Item_MRP': ['sum', 'mean'],
        'Item_Weight': 'sum'
    }).round(2)
    
    outlet_sales.columns = ['總銷售額', '平均銷售額', '總重量']
    
    # 寫入門店銷售數據
    headers = ["門店", "總銷售額", "平均銷售額", "總重量"]
    for col, header in enumerate(headers, 1):
        cell = ws_sales.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row, (outlet, data) in enumerate(outlet_sales.iterrows(), 4):
        ws_sales.cell(row=row, column=1).value = outlet
        for col, value in enumerate(data, 2):
            cell = ws_sales.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
            if col in [2, 3, 4]:  # 金額列
                cell.number_format = '$#,##0.00'
    
    # 4. 商品分析
    ws_products = wb.create_sheet("商品類別分析")
    
    ws_products['A1'] = "商品類別分析報告"
    ws_products['A1'].font = title_font
    ws_products.merge_cells('A1:E1')
    ws_products['A1'].alignment = Alignment(horizontal='center')
    
    product_analysis = df.groupby('Item_Type').agg({
        'Item_MRP': ['sum', 'mean'],
        'Item_Weight': 'sum'
    }).round(2)
    
    product_analysis.columns = ['總銷售額', '平均銷售額', '總重量']
    
    # 寫入商品分析數據
    headers = ["商品類別", "總銷售額", "平均銷售額", "總重量"]
    for col, header in enumerate(headers, 1):
        cell = ws_products.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for row, (item_type, data) in enumerate(product_analysis.iterrows(), 4):
        ws_products.cell(row=row, column=1).value = item_type
        for col, value in enumerate(data, 2):
            cell = ws_products.cell(row=row, column=col)
            cell.value = value
            cell.font = normal_font
            cell.border = border
            if col in [2, 3, 4]:  # 金額列
                cell.number_format = '$#,##0.00'
    
    # 添加總計行
    total_row = len(product_analysis) + 4
    ws_products.cell(row=total_row, column=1).value = "總計"
    ws_products.cell(row=total_row, column=2).value = product_analysis['總銷售額'].sum()
    ws_products.cell(row=total_row, column=3).value = product_analysis['平均銷售額'].mean()
    
    # 設置總計行樣式
    for col in range(1, 5):
        cell = ws_products.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.border = border
    
    # 5. 客戶分析（如果有數據）
    if customer_metrics is not None and segment_analysis is not None:
        ws_customers = wb.create_sheet("客戶群分析")
        
        ws_customers['A1'] = "客戶群分析報告"
        ws_customers['A1'].font = title_font
        ws_customers.merge_cells('A1:E1')
        ws_customers['A1'].alignment = Alignment(horizontal='center')
        
        # 客戶群概況
        headers = ["客戶群", "客戶數量", "平均總消費額", "平均購買次數", 
                  "平均單次消費額", "客戶佔比", "銷售額佔比"]
        for col, header in enumerate(headers, 1):
            cell = ws_customers.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        total_customers = segment_analysis['客戶數量'].sum()
        total_sales = segment_analysis['平均總消費額'].sum()
        
        for row, (segment, data) in enumerate(segment_analysis.iterrows(), 4):
            # 基本數據
            ws_customers.cell(row=row, column=1).value = f"群組 {segment}"
            ws_customers.cell(row=row, column=2).value = data['客戶數量']
            ws_customers.cell(row=row, column=3).value = data['平均總消費額']
            ws_customers.cell(row=row, column=4).value = data['平均購買次數']
            ws_customers.cell(row=row, column=5).value = data['平均單次消費額']
            
            # 計算佔比
            ws_customers.cell(row=row, column=6).value = data['客戶數量'] / total_customers
            ws_customers.cell(row=row, column=7).value = data['平均總消費額'] / total_sales
            
            # 設置格式
            for col in range(1, 8):
                cell = ws_customers.cell(row=row, column=col)
                cell.font = normal_font
                cell.border = border
                if col == 2:  # 數值列
                    cell.number_format = '$#,##0.00'
                elif col in [6, 7]:  # 百分比
                    cell.number_format = '0.00%'
    
    # 調整所有工作表的格式
    for ws in wb.worksheets:
        # 調整列寬
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # 設置行高
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 20
    
    # 保存到BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def add_report_download_section(df, customer_metrics=None, segment_analysis=None):
    """添加報表下載區塊"""
    st.header("📊 下載分析報表")
    
    # 生成報表
    excel_buffer = generate_excel_report(df, customer_metrics, segment_analysis)
    
    # 添加下載按鈕
    current_date = pd.Timestamp.now().strftime("%Y%m%d")
    filename = f"銷售分析報表_{current_date}.xlsx"
    
    st.download_button(
        label="📥 下載完整分析報表 (Excel)",
        data=excel_buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.info("""
    📋 報表內容包括：
    1. 銷售概況
    2. 客戶分析
    3. 商品分析
    """)
        
def main():
    st.title("🏪 智慧商業銷售分析系統")
    
    # 顯示API資訊
    st.sidebar.header("🔌 API資訊")
    st.sidebar.info("""
    本系統提供REST API服務：
    
    1. 分析端點 (POST /analyze)
       - 上傳CSV檔案獲取分析結果
       - 返回JSON格式的分析洞察
       - 需要在請求標頭中提供 X-API-Key
    
    2. 報表端點 (POST /generate-report)
       - 上傳CSV檔案生成報表
       - 支援Excel和ZIP格式
       - 需要在請求標頭中提供 X-API-Key
    
    API文件：http://localhost:8000/docs
    """)
    
    # 上傳資料
    uploaded_file = st.file_uploader("請上傳銷售數據 (CSV格式)", type=['csv'])
    
    if uploaded_file is not None:
        # 讀取並處理數據
        df = load_and_process_data(uploaded_file)
        
        if df is not None:
            # 顯示數據概覽
            st.header("📊 數據概覽")
            st.write(df.head())
            
            # 顯示頂部指標
            show_header_metrics(df)
            
            # 使用 tabs 進行分頁展示
            tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab13, tab14 = st.tabs([
                "📦 商品分析",
                "🏪 商店分析",
                "📈 趨勢預測",
                "🔬 進階分析",
                "🔄 相關性分析",
                "📅 時間序列",
                "💰 價格分析",
                "👥 客戶分析",
                "📊 損益表",
                "👥 客戶收入報表",
                "💰 資產負債表",
                "📈 財務比率分析",
                "📊 營運指標分析"
            ])
              
            with tab2:
                analyze_products(df)
            
            with tab3:
                analyze_stores(df)
            
            with tab4:
                analyze_trends(df)
            
            with tab5:
                perform_advanced_analysis(df)
            
            with tab6:
                perform_correlation_analysis(df)
            
            with tab7:
                perform_time_series_analysis(df)
            
            with tab8:
                perform_price_analysis(df)
            
            with tab9:
                perform_customer_segmentation(df)
            
            with tab10:
                generate_profit_loss_statement(df)
            
            with tab11:
                generate_customer_revenue_report(df)
            
            with tab12:
                generate_balance_sheet(df)
            
            with tab13:
                generate_financial_ratios(df)
            
            with tab14:
                generate_operational_metrics(df)

if __name__ == "__main__":
    main()

